# app.py
from flask_cors import CORS
import pandas as pd
import os
from io import BytesIO
from flask import Flask, request, jsonify, send_file
from flask_sqlalchemy import SQLAlchemy
from werkzeug.security import generate_password_hash, check_password_hash
import openpyxl
from openpyxl.styles import Font
from sqlalchemy import func 

app = Flask(__name__)
CORS(app) # <-- YENİ EKLENEN SATIR

# --- Veritabanı Yapılandırması ---
#basedir = os.path.abspath(os.path.dirname(__file__))
#app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///' + os.path.join(basedir, 'site.db')
#app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

#db = SQLAlchemy(app)

# --- Veritabanı Modelleri ---
class User(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    password_hash = db.Column(db.String(128), nullable=False)
    is_admin = db.Column(db.Boolean, default=False)
    tasks = db.relationship('Task', backref='owner', lazy=True, cascade="all, delete-orphan")

class Task(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(120), nullable=False)
    progress = db.Column(db.Integer, nullable=False, default=0) 
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    subtasks = db.relationship('SubTask', backref='main_task', lazy=True, cascade="all, delete-orphan")

class SubTask(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(120), nullable=False)
    bina_ici = db.Column(db.Integer, default=0)
    bina_disi = db.Column(db.Integer, default=0)
    sozlesme = db.Column(db.Integer, default=0)
    yapilan = db.Column(db.Integer, default=0)
    gereken_gun = db.Column(db.Integer, default=0)
    task_id = db.Column(db.Integer, db.ForeignKey('task.id'), nullable=False)
    
    # app.py dosyasında SubTask modelinin altına ekleyin

class Maliyet(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    is_kalemi = db.Column(db.String(200), nullable=False)
    alt_kalem = db.Column(db.String(300), nullable=True, unique=False)
    isin_bedeli = db.Column(db.Float, default=0.0)
    yapilan_odeme = db.Column(db.Float, default=0.0)
    odeme_sekli = db.Column(db.String(100))
    odeme_tarihi = db.Column(db.String(100)) # Tarihi şimdilik string tutalım
    # YENİ: Soft delete için durum kolonu
    durum = db.Column(db.String(50), default='Aktif') # 'Aktif' veya 'Silindi'

    @property
    def kalan_bakiye(self):
        return self.isin_bedeli - self.yapilan_odeme
        
        
from datetime import datetime
from werkzeug.utils import secure_filename # Dosya adını güvenli hale getirmek için

# ... (diğer modeller) ...

class Gorsel(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    filename = db.Column(db.String(200), nullable=False)
    file_type = db.Column(db.String(50), nullable=False)  # 'image' veya 'video'
    caption = db.Column(db.String(300))
    upload_date = db.Column(db.DateTime, default=datetime.utcnow)
    # Hangi müşteriye ait olduğunu belirtmek için
    customer_username = db.Column(db.String(80), nullable=False)

# --- Yardımcı Fonksiyonlar ---
def calculate_subtask_percentage(yapilan, gereken_gun):
    if gereken_gun is not None and gereken_gun > 0:
        return round((yapilan / gereken_gun) * 100)
    return 0

def get_int_value_from_string(value):
    try:
        if value is None or str(value).strip() == '': return 0
        return int(float(str(value).replace(',', '.')))
    except (ValueError, TypeError): return 0

def normalize_string_for_search(s):
    if s is None: return ""
    cleaned = ''.join(char for char in str(s).strip() if char.isalnum())
    return cleaned.upper().replace('İ', 'I').replace('Ü', 'U').replace('Ö', 'O').replace('Ç', 'C').replace('Ş', 'S').replace('Ğ', 'G')

def find_header_column(sheet, header_name_keywords, search_rows=5):
    for row_idx in range(1, search_rows + 1):
        for col_idx in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(row=row_idx, column=col_idx).value
            if cell_value:
                cleaned_cell_value = normalize_string_for_search(cell_value)
                if any(normalize_string_for_search(keyword) in cleaned_cell_value for keyword in header_name_keywords):
                    return col_idx, row_idx
    return -1, -1

def recalculate_main_task_progress(task_id):
    task = db.session.get(Task, task_id)
    if not task or not task.subtasks:
        if task:
            task.progress = 0
            db.session.commit()
        return
    total_percentage_sum = sum(calculate_subtask_percentage(st.yapilan, st.gereken_gun) for st in task.subtasks)
    new_progress = round(total_percentage_sum / len(task.subtasks))
    if task.progress != new_progress:
        task.progress = new_progress
        db.session.commit()

def clone_tasks_for_new_customer(template_user, new_customer):
    for task_to_clone in template_user.tasks:
        new_task = Task(name=task_to_clone.name, progress=0, owner=new_customer)
        db.session.add(new_task)
        db.session.flush()
        for subtask_to_clone in task_to_clone.subtasks:
            new_subtask = SubTask(name=subtask_to_clone.name, bina_ici=subtask_to_clone.bina_ici, bina_disi=subtask_to_clone.bina_disi, sozlesme=subtask_to_clone.sozlesme, yapilan=0, gereken_gun=subtask_to_clone.gereken_gun, main_task=new_task)
            db.session.add(new_subtask)
    db.session.commit()

# --- Veritabanı ve Başlangıç Verileri ---
with app.app_context():
    db.create_all()
    if not User.query.filter_by(username='admin').first():
        hashed_password = generate_password_hash('123')
        admin_user = User(username='admin', password_hash=hashed_password, is_admin=True)
        db.session.add(admin_user)
        db.session.commit()
    
    admin_user = User.query.filter_by(username='admin').first()
    if not User.query.filter_by(username='Müşteri').first():
        hashed_password = generate_password_hash('123')
        default_customer = User(username='Müşteri', password_hash=hashed_password, is_admin=False)
        db.session.add(default_customer)
        db.session.commit()
        if admin_user.tasks and not default_customer.tasks:
             clone_tasks_for_new_customer(admin_user, default_customer)

    if not admin_user.tasks:
        excel_folder = os.path.join(basedir, 'EXCEL_DOSYALARI')
        excel_file_name = 'Gökhan Tetik İş Takip.xlsx'
        excel_file_path = os.path.join(excel_folder, excel_file_name)
        try:
            workbook = openpyxl.load_workbook(excel_file_path, data_only=True)
            if 'İCMAL' in workbook.sheetnames:
                icmal_sheet = workbook['İCMAL']
                iş_tanımı_col, header_row_idx_icmal = find_header_column(icmal_sheet, ['İŞ TANIMI', 'GÖREV ADI'])
                if iş_tanımı_col != -1:
                    for row_idx in range(header_row_idx_icmal + 1, icmal_sheet.max_row + 1):
                        task_name_value = icmal_sheet.cell(row=row_idx, column=iş_tanımı_col).value
                        if task_name_value and str(task_name_value).strip() != '':
                            task_name = str(task_name_value).strip()
                            if not Task.query.filter_by(name=task_name, user_id=admin_user.id).first():
                                new_task = Task(name=task_name, progress=0, owner=admin_user)
                                db.session.add(new_task)
                    db.session.commit()
            
            admin_tasks = Task.query.filter_by(user_id=admin_user.id).all()
            for sheet in workbook:
                if sheet.title == 'İCMAL': continue
                normalized_sheet_name = normalize_string_for_search(sheet.title)
                main_task_obj = None
                for task in admin_tasks:
                    if normalize_string_for_search(task.name) == normalized_sheet_name:
                        main_task_obj = task
                        break
                if not main_task_obj: continue
                
                # Alt görevlerin zaten var olup olmadığını kontrol et
                if main_task_obj.subtasks: continue

                subtask_name_col, header_row_idx_sub = find_header_column(sheet, ['ALT İŞLER', 'İŞ TANIMI'])
                if subtask_name_col == -1 or header_row_idx_sub == -1: continue

                bina_ici_col, _ = find_header_column(sheet, ['BİNA İÇİ'])
                bina_disi_col, _ = find_header_column(sheet, ['BİNA DIŞI'])
                sozlesme_col, _ = find_header_column(sheet, ['SÖZLEŞME'])
                gereken_gun_col, _ = find_header_column(sheet, ['GEREKEN GÜN'])
                
                for row_idx in range(header_row_idx_sub + 1, sheet.max_row + 1):
                    subtask_name_cell = sheet.cell(row=row_idx, column=subtask_name_col).value
                    if not subtask_name_cell or str(subtask_name_cell).strip() == '': continue
                    
                    new_subtask = SubTask(
                        name=str(subtask_name_cell).strip(),
                        bina_ici=get_int_value_from_string(sheet.cell(row=row_idx, column=bina_ici_col).value if bina_ici_col != -1 else 0),
                        bina_disi=get_int_value_from_string(sheet.cell(row=row_idx, column=bina_disi_col).value if bina_disi_col != -1 else 0),
                        sozlesme=get_int_value_from_string(sheet.cell(row=row_idx, column=sozlesme_col).value if sozlesme_col != -1 else 0),
                        yapilan=0,
                        gereken_gun=get_int_value_from_string(sheet.cell(row=row_idx, column=gereken_gun_col).value if gereken_gun_col != -1 else 0),
                        main_task=main_task_obj
                    )
                    db.session.add(new_subtask)
                db.session.commit()

            default_customer_check = User.query.filter_by(username='Müşteri').first()
            if admin_user.tasks and default_customer_check and not default_customer_check.tasks:
                clone_tasks_for_new_customer(admin_user, default_customer_check)
        except FileNotFoundError:
             print("Excel dosyası 'EXCEL_DOSYALARI' klasöründe bulunamadı.")
        except Exception as e:
            print(f"Excel yükleme sırasında bir hata oluştu: {e}")

# --- API Endpointleri ---

@app.route('/api/login', methods=['POST'])
def login():
    data = request.get_json()
    user = User.query.filter_by(username=data.get('username')).first()
    if user and check_password_hash(user.password_hash, data.get('password')):
        return jsonify({'message': 'Giriş başarılı', 'username': user.username, 'is_admin': user.is_admin}), 200
    else:
        return jsonify({'message': 'Kullanıcı adı veya şifre hatalı'}), 401

@app.route('/api/customers', methods=['GET'])
def get_customers():
    customers = User.query.filter_by(is_admin=False).all()
    customer_data = [{'id': customer.id, 'username': customer.username} for customer in customers]
    return jsonify(customer_data), 200

@app.route('/api/customers', methods=['POST'])
def create_customer():
    data = request.get_json()
    admin_username = data.get('admin_username')
    new_customer_username = data.get('new_customer_username')
    new_customer_password = data.get('new_customer_password')
    admin = User.query.filter_by(username=admin_username, is_admin=True).first()
    if not admin: return jsonify({'message': 'Yetkisiz işlem.'}), 403
    if User.query.filter_by(username=new_customer_username).first(): return jsonify({'message': 'Bu kullanıcı adı zaten mevcut.'}), 409
    hashed_password = generate_password_hash(new_customer_password)
    new_customer = User(username=new_customer_username, password_hash=hashed_password, is_admin=False)
    db.session.add(new_customer)
    db.session.commit()
    template_user = User.query.filter_by(username='admin').first()
    clone_tasks_for_new_customer(template_user, new_customer)
    return jsonify({'message': f"'{new_customer_username}' müşterisi başarıyla oluşturuldu."}), 201

@app.route('/api/tasks/<username>', methods=['GET'])
def get_user_tasks(username):
    user = User.query.filter_by(username=username).first_or_404()
    tasks_data = [{'id': task.id, 'name': task.name, 'progress': task.progress} for task in user.tasks]
    return jsonify(tasks_data), 200

@app.route('/api/tasks', methods=['POST'])
def add_task():
    data = request.get_json()
    admin_username = data.get('username')
    customer_username = data.get('customerUsername')
    admin = User.query.filter_by(username=admin_username, is_admin=True).first()
    if not admin: return jsonify({'message': 'Yetkisiz erişim.'}), 403
    customer = User.query.filter_by(username=customer_username).first()
    if not customer: return jsonify({'message': f"'{customer_username}' müşterisi bulunamadı."}), 404
    task_name = data.get('name')
    if Task.query.filter_by(name=task_name, user_id=customer.id).first(): return jsonify({'message': 'Bu müşteri için bu görev zaten mevcut.'}), 409
    new_task = Task(name=task_name, progress=data.get('progress', 0), owner=customer)
    db.session.add(new_task)
    db.session.commit()
    return jsonify({'message': 'Görev başarıyla eklendi', 'task': {'id': new_task.id, 'name': new_task.name, 'progress': new_task.progress}}), 201

@app.route('/api/tasks/<int:task_id>', methods=['PUT'])
def update_task(task_id):
    task = db.session.get(Task, task_id)
    if not task: return jsonify({'message': 'Görev bulunamadı'}), 404
    data = request.get_json()
    requesting_user = User.query.filter_by(username=data.get('username')).first()
    if not requesting_user or not requesting_user.is_admin: return jsonify({'message': 'Yetkisiz erişim.'}), 403
    if 'name' in data and data['name'] != task.name:
        if Task.query.filter_by(name=data['name'], user_id=task.user_id).first(): return jsonify({'message': 'Bu isimde bir görev zaten mevcut.'}), 409
        task.name = data['name']
    db.session.commit()
    return jsonify({'message': 'Görev başarıyla güncellendi'}), 200

@app.route('/api/tasks/<int:task_id>', methods=['DELETE'])
def delete_task(task_id):
    task = db.session.get(Task, task_id)
    if not task: return jsonify({'message': 'Görev bulunamadı'}), 404
    requesting_user = User.query.filter_by(username=request.headers.get('X-Requested-By-Username')).first()
    if not requesting_user or not requesting_user.is_admin: return jsonify({'message': 'Yetkisiz erişim.'}), 403
    db.session.delete(task)
    db.session.commit()
    return jsonify({'message': 'Görev başarıyla silindi'}), 200

@app.route('/api/subtasks/<int:task_id>', methods=['GET'])
def get_subtasks_for_task(task_id):
    task = db.session.get(Task, task_id)
    if not task: return jsonify({'message': 'Görev bulunamadı'}), 404
    subtasks_data = [{'id': st.id, 'name': st.name, 'binaIci': st.bina_ici, 'binaDisi': st.bina_disi, 'sozlesme': st.sozlesme, 'yapilan': st.yapilan, 'yuzde': calculate_subtask_percentage(st.yapilan, st.gereken_gun), 'gerekenGun': st.gereken_gun} for st in task.subtasks]
    return jsonify(subtasks_data), 200

@app.route('/api/subtasks', methods=['POST'])
def add_subtask():
    data = request.get_json()
    requesting_user = User.query.filter_by(username=data.get('username')).first()
    if not requesting_user or not requesting_user.is_admin: return jsonify({'message': 'Yetkisiz erişim.'}), 403
    task = db.session.get(Task, data.get('task_id'))
    if not task: return jsonify({'message': 'Ana görev bulunamadı'}), 404
    subtask_name = data.get('name')
    if SubTask.query.filter_by(name=subtask_name, task_id=task.id).first(): return jsonify({'message': 'Bu alt görev zaten mevcut.'}), 409
    new_subtask = SubTask(name=subtask_name, bina_ici=data.get('binaIci', 0), bina_disi=data.get('binaDisi', 0), sozlesme=data.get('sozlesme', 0), yapilan=data.get('yapilan', 0), gereken_gun=data.get('gerekenGun', 0), main_task=task)
    db.session.add(new_subtask)
    db.session.commit()
    recalculate_main_task_progress(task.id)
    return jsonify({'message': 'Alt görev başarıyla eklendi'}), 201

@app.route('/api/subtasks/<int:subtask_id>', methods=['PUT'])
def update_subtask(subtask_id):
    subtask = db.session.get(SubTask, subtask_id)
    if not subtask: return jsonify({'message': 'Alt görev bulunamadı'}), 404
    data = request.get_json()
    requesting_user = User.query.filter_by(username=data.get('username')).first()
    if not requesting_user or not requesting_user.is_admin: return jsonify({'message': 'Yetkisiz erişim.'}), 403
    if 'name' in data and data['name'] != subtask.name:
        if SubTask.query.filter_by(name=data['name'], task_id=subtask.task_id).first(): return jsonify({'message': 'Bu isimde bir alt görev zaten mevcut.'}), 409
        subtask.name = data['name']
    subtask.bina_ici = data.get('binaIci', subtask.bina_ici)
    subtask.bina_disi = data.get('binaDisi', subtask.bina_disi)
    subtask.sozlesme = data.get('sozlesme', subtask.sozlesme)
    subtask.yapilan = data.get('yapilan', subtask.yapilan)
    subtask.gereken_gun = data.get('gerekenGun', subtask.gereken_gun)
    db.session.commit()
    recalculate_main_task_progress(subtask.task_id)
    return jsonify({'message': 'Alt görev başarıyla güncellendi'}), 200

@app.route('/api/subtasks/<int:subtask_id>', methods=['DELETE'])
def delete_subtask(subtask_id):
    subtask = db.session.get(SubTask, subtask_id)
    if not subtask: return jsonify({'message': 'Alt görev bulunamadı'}), 404
    requesting_user = User.query.filter_by(username=request.headers.get('X-Requested-By-Username')).first()
    if not requesting_user or not requesting_user.is_admin: return jsonify({'message': 'Yetkisiz erişim.'}), 403
    main_task_id = subtask.task_id
    db.session.delete(subtask)
    db.session.commit()
    recalculate_main_task_progress(main_task_id)
    return jsonify({'message': 'Alt görev başarıyla silindi'}), 200

@app.route('/api/tasks/<int:task_id>/export', methods=['GET'])
def export_task_to_excel(task_id):
    task = db.session.get(Task, task_id)
    if not task:
        return "Görev bulunamadı", 404

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = task.name[:30] 

    headers = ["Alt Görev Adı", "Bina İçi", "Bina Dışı", "Sözleşme", "Yapılan", "Yüzde (%)", "Gereken Gün"]
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True)

    for subtask in task.subtasks:
        percentage = calculate_subtask_percentage(subtask.yapilan, subtask.gereken_gun)
        row_data = [
            subtask.name,
            subtask.bina_ici,
            subtask.bina_disi,
            subtask.sozlesme,
            subtask.yapilan,
            percentage,
            subtask.gereken_gun
        ]
        sheet.append(row_data)

    excel_stream = BytesIO()
    workbook.save(excel_stream)
    excel_stream.seek(0)

    safe_filename = "".join([c for c in task.name if c.isalpha() or c.isdigit() or c in ' -_']).rstrip()
    download_name = f"{safe_filename}_Detaylari.xlsx"

    return send_file(
        excel_stream,
        as_attachment=True,
        download_name=download_name,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

# --- YENİ EKLENEN VE GÜNCELLENEN FONKSİYON ---
@app.route('/api/export-all/<username>', methods=['GET'])
def export_all_to_excel(username):
    user = User.query.filter_by(username=username).first()
    if not user:
        return "Kullanıcı bulunamadı", 404

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = f"{username} Görev Raporu"

    headers = [
        "Ana Görev Adı", "Alt Görev Adı", "Bina İçi", "Bina Dışı",
        "Sözleşme", "Yapılan", "Yüzde (%)", "Gereken Gün"
    ]
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True)

    sorted_tasks = sorted(user.tasks, key=lambda t: t.name)

    for task in sorted_tasks:
        if not task.subtasks:
            row_data = [task.name, "N/A", "N/A", "N/A", "N/A", "N/A", "N/A", "N/A"]
            sheet.append(row_data)
        else:
            sorted_subtasks = sorted(task.subtasks, key=lambda s: s.name)
            for subtask in sorted_subtasks:
                percentage = calculate_subtask_percentage(subtask.yapilan, subtask.gereken_gun)
                row_data = [
                    task.name,
                    subtask.name,
                    subtask.bina_ici,
                    subtask.bina_disi,
                    subtask.sozlesme,
                    subtask.yapilan,
                    percentage,
                    subtask.gereken_gun
                ]
                sheet.append(row_data)

    for col in sheet.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[column].width = adjusted_width

    excel_stream = BytesIO()
    workbook.save(excel_stream)
    excel_stream.seek(0)

    safe_filename = "".join([c for c in user.username if c.isalpha() or c.isdigit() or c in ' -_']).rstrip()
    download_name = f"GoDesign_{safe_filename}_Tum_Gorev_Raporu.xlsx"

    return send_file(
        excel_stream,
        as_attachment=True,
        download_name=download_name,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )



@app.route('/api/costs/summary', methods=['GET'])
def get_costs_summary():
    try:
        # Veritabanındaki 'Aktif' durumdaki kayıtların bedel ve ödeme sütunlarının toplamını al
        query_result = db.session.query(
            func.sum(Maliyet.isin_bedeli),
            func.sum(Maliyet.yapilan_odeme)
        ).filter(Maliyet.durum == 'Aktif').one()

        # Eğer hiç kayıt yoksa toplamlar None gelebilir, bu durumu kontrol et
        total_cost = query_result[0] or 0.0
        total_paid = query_result[1] or 0.0
        remaining_balance = total_cost - total_paid

        return jsonify({
            "total": total_cost,
            "paid": total_paid,
            "remaining": remaining_balance
        })
    except Exception as e:
        print(f"Maliyet özeti alınırken veritabanı hatası: {e}")
        return jsonify({"error": f"Özet verisi alınamadı: {str(e)}"}), 500

# app.py'deki eski maliyet fonksiyonlarını silip bunları ekleyin

@app.route('/api/costs', methods=['GET'])
def get_costs():
    # Sadece 'Aktif' olan kayıtları getir
    costs_from_db = Maliyet.query.filter_by(durum='Aktif').all()
    
    costs_list = []
    for cost in costs_from_db:
        # Değerin None olup olmadığını kontrol et, None ise 0 olarak kabul et
        isin_bedeli = cost.isin_bedeli or 0
        yapilan_odeme = cost.yapilan_odeme or 0
        kalan_bakiye = isin_bedeli - yapilan_odeme # Kalan bakiyeyi burada tekrar hesapla

        costs_list.append({
            'id': cost.id,
            'İş Kalemi': cost.is_kalemi,
            'Alt Kalem / Açıklama': cost.alt_kalem,
            'İşin Bedeli': f"{isin_bedeli:,.2f} TL",
            'Yapılan Ödeme': f"{yapilan_odeme:,.2f} TL",
            'Ödeme Şekli': cost.odeme_sekli,
            'Ödeme Tarihi': cost.odeme_tarihi,
            'Kalan Bakiye': f"{kalan_bakiye:,.2f} TL"
        })
    return jsonify(costs_list)

@app.route('/api/costs', methods=['PUT'])
def update_cost_entry():
    data = request.get_json()
    # Admin kontrolü
    requesting_user = User.query.filter_by(username=data.get('admin_username')).first()
    if not requesting_user or not requesting_user.is_admin:
        return jsonify({'message': 'Yetkisiz işlem.'}), 403

    cost_id = data.get('id')
    updated_data = data.get('updated_data')
    
    cost_entry = db.session.get(Maliyet, cost_id)
    if not cost_entry:
        return jsonify({"error": "Kayıt bulunamadı."}), 404
        
    cost_entry.is_kalemi = updated_data.get('İş Kalemi', cost_entry.is_kalemi)
    cost_entry.alt_kalem = updated_data.get('Alt Kalem / Açıklama', cost_entry.alt_kalem)
    cost_entry.isin_bedeli = clean_currency_for_calc(updated_data.get('İşin Bedeli', cost_entry.isin_bedeli))
    cost_entry.yapilan_odeme = clean_currency_for_calc(updated_data.get('Yapılan Ödeme', cost_entry.yapilan_odeme))
    cost_entry.odeme_sekli = updated_data.get('Ödeme Şekli', cost_entry.odeme_sekli)
    cost_entry.odeme_tarihi = updated_data.get('Ödeme Tarihi', cost_entry.odeme_tarihi)
    
    db.session.commit()
    return jsonify({"message": "Kayıt başarıyla güncellendi."})

@app.route('/api/costs', methods=['DELETE'])
def delete_cost_entry():
    data = request.get_json()
    cost_id = data.get('id')
    
    cost_entry = db.session.get(Maliyet, cost_id)
    if not cost_entry:
        return jsonify({"error": "Kayıt bulunamadı."}), 404
        
    # Satırı veritabanından silmek yerine durumunu 'Silindi' yap (soft delete)
    cost_entry.durum = 'Silindi'
    db.session.commit()
    
    return jsonify({"message": f"'{cost_entry.alt_kalem}' kaydı başarıyla silindi."})

@app.route('/api/costs/export', methods=['GET'])
def export_costs_to_excel():
    customer_name = request.args.get('customer', 'Musteri')
    
    # Veritabanından veriyi al ve DataFrame'e dönüştür
    costs_from_db = Maliyet.query.filter_by(durum='Aktif').all()
    data_for_df = [{
        'İş Kalemi': c.is_kalemi,
        'Alt Kalem / Açıklama': c.alt_kalem,
        'İşin Bedeli': c.isin_bedeli,
        'Yapılan Ödeme': c.yapilan_odeme,
        'Ödeme Şekli': c.odeme_sekli,
        'Ödeme Tarihi': c.odeme_tarihi,
        'Kalan Bakiye': c.kalan_bakiye
    } for c in costs_from_db]
    df = pd.DataFrame(data_for_df)
    
    output = BytesIO()
    df.to_excel(output, index=False, sheet_name='MaliyetRaporu')
    output.seek(0)
    
    download_filename = f"Maliyet_Tablosu_{customer_name}.xlsx"
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=download_filename
    )

# Sayısal değerleri temizlemek için bir yardımcı fonksiyon
# app.py'deki clean_currency_for_calc fonksiyonunun yeni hali
# app.py'deki eski clean_currency_for_calc fonksiyonunu silip yerine bu nihai versiyonu yapıştırın

def clean_currency_for_calc(value):
    """
    '100.000,00 TL' gibi bir string'i 100000.0 gibi bir float sayıya çevirir.
    """
    try:
        if value is None:
            return 0.0
        
        s_value = str(value).strip()
        
        # Binlik ayıracı olan noktaları kaldır
        s_value = s_value.replace('.', '')
        
        # Ondalık ayıracı olan virgülü noktaya çevir
        s_value = s_value.replace(',', '.')
        
        # Sayısal olmayan karakterleri (TL, $, vb.) temizle
        numeric_part = ''
        for char in s_value:
            if char.isdigit() or char == '.':
                numeric_part += char
            else:
                # Sayısal olmayan ilk karakterde dur
                break
        
        return float(numeric_part) if numeric_part else 0.0
    except (ValueError, TypeError):
        return 0.0



# --- TEK SEFERLİK VERİ AKTARMA FONKSİYONU ---

# app.py'deki eski import_costs_from_excel fonksiyonunu silip yerine bunu yapıştırın

# app.py'deki eski import_costs_from_excel fonksiyonunu silip yerine bu "Hata Ayıklama Versiyonunu" yapıştırın

# Temizlenmiş Veri Aktarma Fonksiyonu
# app.py'deki import_costs_from_excel fonksiyonunun yeni hali
def import_costs_from_excel():
    with app.app_context():
        if Maliyet.query.first() is None:
            try:
                file_path = os.path.join(basedir, 'DATA', 'maliyet.xlsx')
                df = pd.read_excel(file_path).fillna('') # Boş hücreleri en baştan boş metin yap
                print("Excel dosyası okundu, veritabanına aktarılıyor...")

                for index, row in df.iterrows():
                    bedel = clean_currency_for_calc(row.get('İşin Bedeli'))
                    odeme = clean_currency_for_calc(row.get('Yapılan Ödeme'))

                    yeni_maliyet = Maliyet(
                        is_kalemi=row.get('İş Kalemi', ''),
                        alt_kalem=row.get('Alt Kalem / Açıklama', ''),
                        isin_bedeli=bedel,
                        yapilan_odeme=odeme,
                        odeme_sekli=str(row.get('Ödeme Şekli', '')),
                        odeme_tarihi=str(row.get('Ödeme Tarihi', '')),
                        durum='Aktif'
                    )
                    db.session.add(yeni_maliyet)

                db.session.commit()
                print("Maliyet verileri Excel'den veritabanına başarıyla aktarıldı.")

            except FileNotFoundError:
                print("'maliyet.xlsx' dosyası bulunamadı. Maliyet verisi aktarılamadı.")
            except Exception as e:
                print(f"Maliyet verisi aktarılırken bir hata oluştu: {e}")
                db.session.rollback()
    
    # Sunucu başlamadan önce veritabanı tablolarını oluştur ve veriyi aktar
with app.app_context():
    db.create_all()
    import_costs_from_excel()
    
    # app.py dosyasının sonuna ekleyin

# --- GÖRSEL GALERİSİ API ENDPOINT'LERİ ---

UPLOAD_FOLDER = os.path.join(basedir, 'static', 'uploads')
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif', 'mp4', 'mov', 'avi'}
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)

def allowed_file(filename):
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

@app.route('/api/gorseller/<customer_username>', methods=['GET'])
def get_gorseller(customer_username):
    gorseller_from_db = Gorsel.query.filter_by(customer_username=customer_username).order_by(Gorsel.upload_date.desc()).all()
    gorseller_list = [{
        'id': g.id,
        'filename': g.filename,
        'file_type': g.file_type,
        'caption': g.caption,
        'url': f'/static/uploads/{g.filename}' # <-- DEĞİŞİKLİK BURADA
    } for g in gorseller_from_db]
    return jsonify(gorseller_list)

@app.route('/api/gorseller', methods=['POST'])
def upload_gorsel():
    admin_user = User.query.filter_by(username=request.form.get('admin_username'), is_admin=True).first()
    if not admin_user:
        return jsonify({'message': 'Yetkisiz işlem.'}), 403

    if 'file' not in request.files:
        return jsonify({'message': 'Dosya seçilmedi.'}), 400
    
    file = request.files['file']
    customer_username = request.form.get('customer_username')
    caption = request.form.get('caption')

    if file.filename == '' or not customer_username:
        return jsonify({'message': 'Eksik bilgi (dosya veya müşteri seçilmedi).'}), 400

    if file and allowed_file(file.filename):
        original_filename = secure_filename(file.filename)
        # Dosya adını benzersiz yapmak için zaman damgası ekleyelim
        unique_filename = f"{datetime.now().strftime('%Y%m%d%H%M%S')}_{original_filename}"
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], unique_filename)
        file.save(file_path)

        file_extension = original_filename.rsplit('.', 1)[1].lower()
        file_type = 'image' if file_extension in ['png', 'jpg', 'jpeg', 'gif'] else 'video'

        yeni_gorsel = Gorsel(
            filename=unique_filename,
            file_type=file_type,
            caption=caption,
            customer_username=customer_username
        )
        db.session.add(yeni_gorsel)
        db.session.commit()
        
        return jsonify({'message': 'Dosya başarıyla yüklendi.'}), 201

    return jsonify({'message': 'İzin verilmeyen dosya türü.'}), 400

@app.route('/api/gorseller/<int:gorsel_id>', methods=['DELETE'])
def delete_gorsel(gorsel_id):
    admin_user = User.query.filter_by(username=request.json.get('admin_username'), is_admin=True).first()
    if not admin_user:
        return jsonify({'message': 'Yetkisiz işlem.'}), 403
    
    gorsel = db.session.get(Gorsel, gorsel_id)
    if not gorsel:
        return jsonify({'message': 'Görsel bulunamadı.'}), 404
        
    try:
        # Önce fiziksel dosyayı sil
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], gorsel.filename)
        if os.path.exists(file_path):
            os.remove(file_path)
            
        # Sonra veritabanı kaydını sil
        db.session.delete(gorsel)
        db.session.commit()
        return jsonify({'message': 'Görsel başarıyla silindi.'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'message': f'Silme sırasında hata oluştu: {e}'}), 500
        
        # app.py dosyasının sonuna ekleyin

@app.route('/api/costs', methods=['POST'])
def add_cost_entry():
    data = request.get_json()
    
    # Güvenlik: Sadece adminlerin yeni kayıt ekleyebildiğinden emin ol
    requesting_user = User.query.filter_by(username=data.get('admin_username')).first()
    if not requesting_user or not requesting_user.is_admin:
        return jsonify({'message': 'Yetkisiz işlem.'}), 403
    
    # Gerekli alanların varlığını kontrol et
    if not data.get('alt_kalem'):
        return jsonify({'message': 'Alt Kalem / Açıklama alanı boş bırakılamaz.'}), 400

    try:
        # Gelen sayısal verileri güvenli bir şekilde işle
        bedel = clean_currency_for_calc(data.get('isin_bedeli', 0))
        odeme = clean_currency_for_calc(data.get('yapilan_odeme', 0))

        yeni_maliyet = Maliyet(
            is_kalemi=data.get('is_kalemi', ''),
            alt_kalem=data.get('alt_kalem'),
            isin_bedeli=bedel,
            yapilan_odeme=odeme,
            odeme_sekli=data.get('odeme_sekli', ''),
            odeme_tarihi=data.get('odeme_tarihi', ''),
            durum='Aktif'  # Yeni kayıt her zaman Aktif başlar
        )
        
        db.session.add(yeni_maliyet)
        db.session.commit()

        return jsonify({'message': 'Yeni maliyet kaydı başarıyla eklendi.'}), 201

    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f"Kayıt eklenirken bir hata oluştu: {str(e)}"}), 500
        
#if __name__ == '__main__':
# port = int(os.environ.get("PORT", 5000))
#app.run(host='0.0.0.0', port=port)
